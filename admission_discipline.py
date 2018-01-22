# -*- coding: utf-8 -*-
import openpyxl
import xml.dom.minidom  

def open_excel(file):
    try:
        data = xlrd.open_workbook(file)  #xlrd 操作excel的外部库
        return data
    except IOError as e:
        print(str(e))

def excel_table_byindex(file, colnnameindex=0, by_index=0): 

    wb = openpyxl.load_workbook('--入院相关病历--.xlsx')
    table = wb.get_sheet_by_name('Sheet1')
    nrows = table.max_row
    column = table.max_column

    uid_last =0
    for nrow in range(2, nrows):        #遍历每一行 使用openpyxl 数据从row=2,column=1开始
        if nrow == 0:
            continue      
        patient_info = '病人基本信息'
        medical_history = '病史'
        check_up = '体格检查'
        diagnose = '诊断'
        #print(check_up.getdefaultencoding())
        
        uid = table.cell(row=nrow, column=1).value   #取值..第一列      
        if uid!=uid_last:                            #表格id 发生改变 则创建新用户xml
            uid_last = uid
            print(uid_last) 
            doc = xml.dom.minidom.Document()    #打开xml对象          
            xmain = doc.createElement('main_{}'.format(uid))  
            doc.appendChild(xmain)   
            item1 = doc.createElement(patient_info)
            item2 = doc.createElement(medical_history)
            item3 = doc.createElement(check_up)
            item4 = doc.createElement(diagnose)
            for  nrow in range(2,nrows):        #对新用户进行表格遍历 同一个id对应多行信息抽取
                if table.cell(row=nrow,column=1).value == uid_last:
                    print(table.cell(row=nrow,column=6).value)
                    if patient_info in table.cell(row=nrow,column=6).value:
                        p_info = str(table.cell(row=nrow,column=7).value)+':'+str(table.cell(row=nrow,column=8).value)
                        p_info = doc.createTextNode(p_info)
                        item1.appendChild(p_info)

                    if medical_history in table.cell(row=nrow,column=6).value:
                        p_info = str(table.cell(row=nrow,column=7).value)+':'+str(table.cell(row=nrow,column=8).value)
                        p_info = doc.createTextNode(p_info)                        
                        item2.appendChild(p_info)

                    if check_up in table.cell(row=nrow,column=6).value:
                        p_info = str(table.cell(row=nrow,column=7).value)+':'+str(table.cell(row=nrow,column=8).value)
                        p_info = doc.createTextNode(p_info)
                        item3.appendChild(p_info)

                    if diagnose in table.cell(row=nrow,column=6).value:
                        p_info = str(table.cell(row=nrow,column=7).value)+':'+str(table.cell(row=nrow,column=8).value)
                        print(p_info)
                        p_info = doc.createTextNode(p_info)
                        item4.appendChild(p_info)

                    xmain.appendChild(item1)
                    xmain.appendChild(item2)
                    xmain.appendChild(item3)
                    xmain.appendChild(item4)
                    fp= open('入院记录 xml\\patient_{}.xml'.format(uid),'w')
                    doc.writexml(fp,indent='\t',addindent='\t',newl='\n',encoding='utf-8')


 
excel_table_byindex('--入院相关病历--.xlsx')   
